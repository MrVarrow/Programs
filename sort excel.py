import os.path
import pandas as pd
import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import messagebox

class Interface:
    def __init__(self):
        self.root =Tk()
        self.root.title('Sort excel')
        self.root.geometry('800x600')
        self.root.resizable(False, False)
        self.sheet_list = None

    def layout(self):

        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_columnconfigure(2, weight=1)

        heading = Label(self.root, text="Sort your data in excel", font=('Arial', 20), bg="gray")
        heading.grid(row=0, columnspan=3, column=0, sticky=EW, ipady=20)

        pick_file_lbl = Label(self.root, text="Pick your excel file:", font=('Arial', 15))
        pick_file_lbl.grid(row=1, column=0, columnspan=2, pady=15, sticky=EW)

        pick_sheet_lbl = Label(self.root, text="Pick sheet:", font=('Arial', 15))
        pick_sheet_lbl.grid(row=3, column=0, columnspan=2, pady=15, sticky=EW)

        pick_column_sort_lbl = Label(self.root, text="Pick column/s to sort by:", font=('Arial', 15))
        pick_column_sort_lbl.grid(row=5, column=0, columnspan=2, pady=20, sticky=EW)

        self.your_file_path_lbl = Label(self.root, text="Your sorted file path: EMPTY", font=('Arial', 12), bg="light gray",
                                   width=50, height=2, wraplength=450)
        self.your_file_path_lbl.grid(row=8, column=0, pady=5)

        self.picked_file_lbl = Label(self.root, text="EMPTY", font=('Arial', 12), bg="light gray", width=50, height=2,
                                wraplength=450)
        self.picked_file_lbl.grid(row=2, column=0, pady=10, ipady=10, padx=50)

        browse_button =Button(self.root, text="Browse", font=('Arial', 15), bg="gray", command=self.browse_files)
        browse_button.grid(row=2, column=1, sticky=W, ipady=5, ipadx=20)

        submit_file_button = Button(self.root, text="Submit", font=('Arial', 15), bg="gray", command=self.submit_file)
        submit_file_button.grid(row=2, column=2, sticky=E, ipady=5, ipadx=10, padx=20)

        sumbmit_sheet_button = Button(self.root, text="Submit", font=('Arial', 15), bg="gray", command=self.sumbit_sheet)
        sumbmit_sheet_button.grid(row=4, column=1, sticky=W, ipady=5, ipadx=10, padx=20)

        sort_button = Button(self.root, text="Sort", font=('Arial', 15), bg="gray", command=self.sort)
        sort_button.grid(row=6, rowspan=2, column=1, ipady=5, ipadx=40)

        exit_button = Button(self.root, text="Exit", font=('Arial', 15), command=self.root.destroy, bg="gray")
        exit_button.grid(row=9, column=2, sticky=SE, padx=5, pady=5)

        self.sheet_list_widget = ttk.Combobox(self.root, font=("Arial", 15))
        self.sheet_list_widget.grid(row=4, column=0, ipadx=30, ipady=5)

        self.column1_list = ttk.Combobox(self.root, font=("Arial", 15))
        self.column1_list.grid(row=6, sticky=W, pady=10, padx=20)

        self.column2_list = ttk.Combobox(self.root, font=("Arial", 15))
        self.column2_list.grid(row=6, sticky=E, pady=10, padx=20)

        self.column3_list = ttk.Combobox(self.root, font=("Arial", 15))
        self.column3_list.grid(row=7, sticky=W, pady=10, padx=20)

        self.column4_list = ttk.Combobox(self.root, font=("Arial", 15))
        self.column4_list.grid(row=7, sticky=E, pady=10, padx=20)

    def browse_files(self):
        self.filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                                   filetypes=(("Excel file", "*.xlsx*"), ("all files", "*.*")))
        self.picked_file_lbl.configure(text="Picked file: " + self.filename)

    def submit_file(self):
        filepath = self.filename
        load = load_workbook(filepath, read_only=True, keep_links=False)
        self.sheet_list = load.sheetnames
        self.sheet_list_widget.configure(values=self.sheet_list)

    def sumbit_sheet(self):
        self.sheet = self.sheet_list_widget.get()
        self.file = self.filename.split('/')[-1]
        self.excel = pd.ExcelFile(self.file)
        self.x = self.excel.parse(self.sheet)
        self.df = pd.read_excel(self.file, sheet_name=self.sheet)
        self.columns = list(self.df.columns)
        self.column1_list.configure(values=self.columns)
        self.column2_list.configure(values=self.columns)
        self.column3_list.configure(values=self.columns)
        self.column4_list.configure(values=self.columns)

    def sort(self):
        col1 = self.column1_list.get()
        col2 = self.column2_list.get()
        col3 = self.column3_list.get()
        col4 = self.column4_list.get()

        if col1 == "" and col2 == "" and col3 == "" and col4 == "":
            messagebox.showinfo(title="Error", message="Please select at least one column to sort data.")
        elif col1 in self.columns and col2 == "":
            self.x = self.df.sort_values(by=self.column1_list.get())
            self.write_to_excel()

        elif col1 in self.columns and col2 in self.columns and col3 == "":
            self.x = self.df.sort_values(by=[col1, col2], ascending=[True, True])
            self.write_to_excel()

        elif col1 in self.columns and col2 in self.columns and col3 in self.columns and col4 == "":
            self.x = self.df.sort_values(by=self.column1_list.get())
            self.write_to_excel()

        elif col1 in self.columns and col2 in self.columns and col3 in self.columns and col4 in self.columns:
            self.x = self.df.sort_values(by=self.column1_list.get())
            self.write_to_excel()

    def write_to_excel(self):
        beta_name = self.file.split(".")[0]
        new_file_name = beta_name + "-sorted.xlsx"
        writer = pd.ExcelWriter(new_file_name)

        for i in range(len(self.columns)):
            self.x.to_excel(writer, sheet_name=self.sheet, columns=[self.columns[i]], index=False, startcol=i)

        writer._save()
        self.your_file_path_lbl.configure(text=os.path.abspath(new_file_name))


def main():
    run = Interface()
    run.layout()
    mainloop()

if __name__ == '__main__':
    main()